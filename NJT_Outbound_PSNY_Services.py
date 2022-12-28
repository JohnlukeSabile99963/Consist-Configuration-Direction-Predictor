import csv
import pandas as pd
import random
from openpyxl import load_workbook
import seaborn as sns
import matplotlib.pyplot as plt
# from sklearn.model_selection import train_test_split
# from sklearn.metrics import confusion_matrix
# from sklearn.metrics import classification_report
# from sklearn.ensemble import HistGradientBoostingRegressor
# from sklearn import svm
# from sklearn import tree
# from IPython.core.pylabtools import figsize
# from mlxtend.plotting import plot_decision_regions


def convert_xlsx_to_csv(excel_file, csv_file):
    NJT_OB_PSNY_Departures_xlsx = load_workbook(filename=excel_file)
    xlsx_sheet = NJT_OB_PSNY_Departures_xlsx.active
    Departure_Info = []

    # Read all Departure Info from xlsx
    for item in xlsx_sheet.iter_rows(values_only=True):
        Departure_Info.append(list(item))

    # Convert & Write to CSV
    with open(csv_file, 'w') as csv_file_upd:
        converter = csv.writer(csv_file_upd, delimiter=',')
        for line in Departure_Info:
            converter.writerow(line)


if __name__ == "__main__":
    convert_xlsx_to_csv("NJT_Outbound_PSNY_Departures.xlsx", "NJT_Outbound_PSNY_Departures.csv")

new_csv = pd.read_csv("NJT_Outbound_PSNY_Departures.csv")


# Line Key:
# M&E = Morris & Essex Line, or Morristown Line, featuring Gladstone Branch.
#  Termini include Summit, Dover, Peapack, and Gladstone.
# MOBO = Montclair-Boonton Line. Trains terminate/originate at Montclair State University.
# NJCL = North Jersey Coast Line. Termini include Rahway, South Amboy, Long Branch, and Bay Head.
# NEC = Northeast Corridor. Termini include Rahway, Jersey Avenue, and Trenton.
# RVL = Raritan Valley Line. Termini include Raritan and High Bridge.
#  Most trains that serve this line originate/terminate at Newark Penn Station.
#  On weekdays, eight round trips operate directly to/from Penn Station New York, with a stop at Secaucus Junction UL.
#  These trips operate off-peak only and either skip only one station (Garwood) or stop at all stations en route.
#  On weekends and holidays, all RVL trains operate between Newark Penn and Raritan only.


new_csv['Configuration Direction'] = pd.Series(random.choices(['R', 'W'],
                                                              weights=[1, 1],
                                                              k=len(new_csv)))
new_csv.head()


# Velocity and Acceleration are added so that we can further compare departure and travel time data.
# Average Velocity = Distance Covered (Miles) / Expected Travel Time (Mins)
# Average Accleration = Average Velocity / Expeced Travel Time (Mins)

new_csv['Expected Average Speed En Route (mph)'] = 60 * (pd.Series(new_csv['Distance Covered (Miles)']
                                                         / new_csv['Expected Travel Time En Route (Mins)']))
new_csv.head()

new_csv['Expected Average Acceleration En Route'] = 9.81 * (pd.Series(new_csv['Expected Average Speed En Route (mph)']
                                                            / new_csv['Expected Travel Time En Route (Mins)']))
new_csv.head()


# Raritan Valley Line Departures will always have their consist in the "Right" Direction,
# so we must by default set all RVL services to 'R' for the Config. Direction class.
# RVL Departures occupy Rows 260-267; Column J is the Config. Direction Class.

new_csv.loc[258, 'Configuration Direction'] = 'R'
new_csv.loc[259, 'Configuration Direction'] = 'R'
new_csv.loc[260, 'Configuration Direction'] = 'R'
new_csv.loc[261, 'Configuration Direction'] = 'R'
new_csv.loc[262, 'Configuration Direction'] = 'R'
new_csv.loc[263, 'Configuration Direction'] = 'R'
new_csv.loc[264, 'Configuration Direction'] = 'R'
new_csv.loc[265, 'Configuration Direction'] = 'R'

print(new_csv)


# New xlsx file now has the configuration direction with "random" direction outputs 'R' and 'W'
new_csv.to_excel(r'C:\Users\10000\PycharmProjects\NJT_Outbound_PSNY_Departures_CSV_Conv'
                 r'\NJT_Updated_Departure_Info_csv.xlsx', index=False)


# Convert Updated xlsx to Updated csv
def updated_xlsx_to_updated_csv(updated_xlsx, updtd_csv):
    Updated_Departures_xlsx = load_workbook(filename=updated_xlsx)
    Updated_xlsx_sheet = Updated_Departures_xlsx.active
    Updated_Departure_Info = []

    # Read through all updated info including 'Configuration Direction' Class
    for service in Updated_xlsx_sheet.iter_rows(values_only=True):
        Updated_Departure_Info.append(list(service))

    # Convert to Updated CSV
    with open(updtd_csv, 'w') as Updated_Departures_csv:
        converter_for_updated = csv.writer(Updated_Departures_csv, delimiter=',')
        for line in Updated_Departure_Info:
            converter_for_updated.writerow(line)


if __name__ == "__main__":
    updated_xlsx_to_updated_csv("NJT_Updated_Departure_Info_csv.xlsx", "NJT_Updated_Outbound_PSNY_Departures_Info.csv")

Updated_Departure_Info = pd.read_csv("NJT_Updated_Outbound_PSNY_Departures_Info.csv")
print(Updated_Departure_Info)

# Visualize Departure Data with Pairplot:


plt.figure(1)
sns.countplot(data=Updated_Departure_Info, x="Configuration Direction", hue="Line")

# Lines:
# M&E = Morris & Essex Line, or Morristown Line. Gladstone Branch Services included.
#  Termini include Summit, Dover, Peapack, and Gladstone.
# MOBO = Montclair-Boonton Line. Terminus is Montclair State University, or MSU.
# NJCL = North Jersey Coast Line. Termini include Rahway, South Amboy, Long Branch, and Bay Head.
# NEC = Northeast Corridor Line. Termini include Rahway, Jersey Avenue, and Trenton Transit Center.

plt.title("NJ Transit Outbound Departures from Penn Station")
plt.xlabel("Configuration Direction Class: R for 'Right' Consist Direction, W for 'Wrong' Consist Direction")
plt.ylabel("Trains Per Line")
plt.show()


plt.figure(2)
sns.countplot(data=Updated_Departure_Info, x="Configuration Direction", hue="Destination")
plt.title("NJ Transit Outbound Departures from Penn Station")
plt.xlabel("Configuration Direction Class: R for 'Right' Consist Direction, W for 'Wrong' Consist Direction")
plt.ylabel("Trains Per Destination from PSNY")
plt.show()


plt.figure(3)
sns.countplot(data=Updated_Departure_Info, x="Configuration Direction", hue="Service Day")
plt.title("NJ Transit Outbound Departures from Penn Station")
plt.xlabel("Configuration Direction Class: R for 'Right' Consist Direction, W for 'Wrong' Consist Direction")
plt.ylabel("Trains Per Service Period")
plt.show()


plt.figure(4)
sns.countplot(data=Updated_Departure_Info, x="Configuration Direction", hue="Departure Timing")
plt.title("NJ Transit Outbound Departures from Penn Station")
plt.xlabel("Configuration Direction Class: R for 'Right' Consist Direction, W for 'Wrong' Consist Direction")
plt.ylabel("Trains Per Departure Time")
plt.show()


plt.figure(5)
sns.countplot(data=Updated_Departure_Info, x="Configuration Direction", hue="Stations Stopped at En Route")
plt.title("NJ Transit Outbound Departures from Penn Station")
plt.xlabel("Configuration Direction Class: R for 'Right' Consist Direction, W for 'Wrong' Consist Direction")
plt.ylabel("Trains Per Stopped Station Count")
plt.show()


plt.figure(6)
sns.countplot(data=Updated_Departure_Info, x="Configuration Direction", hue="Service Class")
# If skipped stations < 3, Service Class = Local.
# Semi-Local and Express services may vary.
plt.title("NJ Transit Outbound Departures from Penn Station")
plt.xlabel("Configuration Direction Class: R for 'Right' Consist Direction, W for 'Wrong' Consist Direction")
plt.ylabel("Trains Per Service Class")
plt.show()

plt.figure(7)
sns.countplot(data=Updated_Departure_Info, x="Configuration Direction",
              hue="Distance Covered (Miles)")
# Distance Covered in intervals of 5.
plt.title("NJ Transit Outbound Departures from Penn Station")
plt.xlabel("Configuration Direction Class: R for 'Right' Consist Direction, W for 'Wrong' Consist Direction")
plt.ylabel("Trains Per Service Class")
plt.show()
